import os
from openpyxl import Workbook, load_workbook
from modelos.Carrera import Carrera
from dao.carrera_dao_base import CarreraDaoBase

class CarreraDAOXLSX(CarreraDaoBase):
    def __init__(self, filepath: str):
        self.filepath = filepath
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "Carreras"
            ws.append(["id", "nombre", "facultad"])
            wb.save(filepath)

    def add_carrera(self, carrera: Carrera):
        carreras = self.get_all_carreras()
        carrera_id = len(carreras) + 1

        wb = load_workbook(self.filepath)
        ws = wb["Carreras"]
        ws.append([carrera_id, carrera.nombre, carrera.facultad])
        wb.save(self.filepath)

    def get_all_carreras(self) -> list[Carrera]:
        wb = load_workbook(self.filepath)
        ws = wb["Carreras"]
        carreras = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            id_, nombre, facultad = row
            carreras.append(Carrera(int(id_), nombre, facultad))

        return carreras